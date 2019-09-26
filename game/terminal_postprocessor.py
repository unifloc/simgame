"""
Модуль для итогового преобразования данных в требуемый формат МЭР и ТР для выдачи студентам

Кобзарь О., Носачев А. 22.09.2019
"""
import pandas as pd
import os
import random
import subprocess

#pd.set_option('precision', 3)
#pd.set_eng_float_format(accuracy=3)
# TODO сделать форматирование чисто из питона, без шаблонов
# TODO убрать излишнюю точность
class TrueReporMaker():
    def __init__(self):
        self.team_name = None
        self.current_dir = None
        self.path_to_team_directory = None
        self.df = None
        self.wellnames = None
        self.report = None


    def make_month_report(self, team_name):
        self.current_dir = os.getcwd()
        self.team_name = team_name
        self.path_to_team_directory = self.current_dir + "/resultspace/" + team_name + "/"
        self.df = pd.read_csv(self.path_to_team_directory + "sim_result.csv")
        names=self.df.columns[1:]
        wellnames=[]
        for n in names:
            a=n.split(':')
            b=a[1]
            if b in wellnames:
                break
            else:
                wellnames.append(b)
        self.wellnames = wellnames

        time = pd.to_datetime(self.df.time)
        self.report=pd.DataFrame()
        self.report['Time']=time
        self.report.reindex(self.report['Time'])


        for name in self.wellnames:
            self.report[name + '. Дебит по нефти, м3/сут']=self.df['WOPR:'+name]
            self.report[name + '. Дебит по жидкости, м3/сут']=self.df['WLPR:'+name]
            self.report[name + '. Дебит по газу, м3/сут']=self.df['WGPR:'+name]
            self.report[name + '. Закачка воды, м3/сут']=self.df['WWIR:'+name]

        self.report.to_excel(self.path_to_team_directory + "МЭР {} сырой.xlsx".format(team_name),
                             sheet_name='МЭР'.format(team_name))
        self.report = self.report.resample('M', on='Time').last()
        self.report = self.report.interpolate()
        del self.report['Time']
        self.report.to_excel(self.path_to_team_directory + "МЭР {}.xlsx".format(team_name),
                             sheet_name='МЭР'.format(team_name))


    def append_df_to_excel(self, filename, df, sheet_name='Sheet1', startrow=None, truncate_sheet=False, **to_excel_kwargs):
        from openpyxl import load_workbook
        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        writer = pd.ExcelWriter(filename, engine='openpyxl')

        try:
            # try to open an existing workbook
            writer.book = load_workbook(filename)

            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row

            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)

            # copy existing sheets
            writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
        except FileNotFoundError:
            # file does not exist yet, we will create it
            pass

        if startrow is None:
            startrow = 0
        # write out the new sheet
        df.style.set_precision(precision=3)
        df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

        # save the workbook
        writer.save()
    def calc_watercut_d(self, wellname):
        q_liq_m3day = self.df['WLPR:'+wellname].tail(1).mean()
        if q_liq_m3day != 0:
            watercut_d = self.df['WWPR:'+wellname].tail(1).mean()/self.df['WLPR:'+wellname].tail(1).mean()
            return watercut_d
        else:
            return 0
    def make_tech_regime(self, team_name):
        d = []
        for name in self.wellnames:
            watercut_d = self.calc_watercut_d(name)
            d.append({'01_НГДУ': team_name,
            '02_Месторождение': 'RIENM1',
            '03_Скважина':name,
            '04_Тип_скважины':'верт',
            '05_Пласт':"м1",
            '06_Dскв': 200,
            '07_Dнкт':73,
            '08_Н_вдп':2500, # TODO индивидуально для каждой команды, исправить
            '09_Удл':0,
            '10_Ндин':random.randint(200,700),
            '11_СЭ':'ESP',  # TODO может быть фонтан, справить
            '12_Рзаб': "", #self.df['WBHP:'+name].tail(1).mean(), # TODO выдавать ли значения, пока нет TM?
            '13_Qнефти':self.df['WOPR:'+name].tail(1).mean(), # TODO посмотреть значения, брать ли последние?
            '14_Qжидк':self.df['WLPR:'+name].tail(1).mean(),
            '15_Обводненность': watercut_d,
            '16_Pзатр':random.randint(11,15),
            '17_ГФ': self.df['WGOR:'+name].tail(1).mean(),
            '18_Тпл': 104, # TODO поставить реальные значения здесь и далее
            '19_Пл-ть_нефти': 0.85,
            '20_Пл-ть_воды': 1})
        df1=pd.DataFrame(d)
        empty_regime_path = self.current_dir + "/resultspace/ТР_шаблон.xlsx"
        full_regime_path = self.path_to_team_directory + "/ТР {}.xlsx".format(team_name)
        subprocess.call(["cp", '-f', empty_regime_path, full_regime_path])  # TODO исправить ошибку с путями и копированием шаблона
        self.append_df_to_excel(full_regime_path, df1, sheet_name='TR',
                           startrow=9 + 1, startcol=1, index=False, header=False)



#team_names = ['ФОН', "FlexOil"]
def run_terminal_postprocessor(team_names):
    for this_team_name in team_names:
        print("---Формирование итоговых файлов (МЭР и ТР) для команды " + this_team_name + ".---" )
        object_team = TrueReporMaker()
        object_team.make_month_report(this_team_name)
        object_team.make_tech_regime(this_team_name)

from ecl.summary import EclSum
import pandas as pd
import random


path_to_sim_dir = "./workspace/"
model_data_file_name = "spe1.DATA"

file_name = path_to_sim_dir + model_data_file_name
summary = EclSum(file_name)
dates = summary.dates

shape = len(dates)
result_df = pd.DataFrame({"test": list(range(shape))})

nedeed_keys = ["WOPR:*", "WWPR:*", "WLPR:*", "WGPR:*", "WWIR:*" ,
               "WGOR:*", "WBHP:*",
               "WOPT:*", "WWPT:*", "WLPT:*", "WGPT:*", "WWIT:*",
               "FOPT", "FWPT", "FLPT", "FGPT", "FWIT"]

for i in nedeed_keys:
    keys_by_wells = summary.keys(i)
    for j in keys_by_wells:
        this_parameter_values = summary.numpy_vector(j)
        one_parameter_df = pd.DataFrame({j: this_parameter_values})
        result_df = result_df.join(one_parameter_df)

time_parameter_column = pd.DataFrame({"time": dates})
result_df = result_df.join(time_parameter_column)
result_df = result_df.set_index("time")
del result_df['test']

result_df.to_csv("sim_result.csv")

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,truncate_sheet=False, **to_excel_kwargs):
    from openpyxl import load_workbook
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
        
    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
    
    
# Export to csv
def export_to_csv(path, name):
    
    df=pd.read_csv(path + f"/resultspace/{name}/sim_result.csv")
    names = df.columns[1:]
    wellnames = []
    for n in names:
        a=n.split(':')
        b=a[1]
        if b in wellnames:
            break
        else:
            wellnames.append(b)
    d = []
    for name in wellnames:
        d.append({'01_НГДУ':'RIENM Corp',    
                  '02_Месторождение': 'RIENM1',
                  '03_Скважина':name,         
                  '04_Тип_скважины':'верт',
                  '05_Пласт':"м1",         
                  '06_Dскв': 200,
                  '07_Dнкт':73,
                  '08_Н_вдп':2500,
                  '09_Удл':0,          
                  '10_Ндин':random.randint(200,500),
                  '11_СЭ':'ESP',
                  '12_Рзаб':df['WBHP:'+name].tail(20).mean(),        
                  '13_Qнефти':df['WOPR:'+name].tail(20).mean(),
                  '14_Qжидк':df['WLPR:'+name].tail(20).mean(),
                  '15_Обводненность':df['WWPR:'+name].tail(20).mean()/1+df['WLPR:'+name].tail(20).mean(),
                  '16_Pзатр':15,
                  '17_ГФ':60,
                  '18_Тпл':104,
                  '19_Пл-ть_нефти':0.85,
                  '20_Пл-ть_воды':1 })
        df1=pd.DataFrame(d)
        append_df_to_excel("201910_TR_1.xlsx", df1,sheet_name='TR', 
                            startrow=0, startcol=0, index=False, header=False)
        

from ecl.summary import EclSum
import pandas as pd
import random

file_name = "./workspace/SPE1.DATA"
summary = EclSum(file_name)
dates = summary.dates

nedeed_keys = ["WOPR:*", "WWPR:*", "WLPR:*", "WGPR:*", "WWIR:*" ,
                    "WGOR:*", "WBHP:*",
                    "WOPT:*", "WWPT:*", "WLPT:*", "WGPT:*", "WWIT:*",
                    "FOPT", "FWPT", "FLPT", "FGPT", "FWIT"]

shape = len(dates)
result_df = pd.DataFrame({"test": list(range(shape))})

for i in nedeed_keys:
    keys_by_wells = summary.keys(i)
    for j in keys_by_wells:
        this_parameter_values = summary.numpy_vector(j)
        one_parameter_df = pd.DataFrame({j: this_parameter_values})
        result_df = result_df.join(one_parameter_df)

time_parameter_column = pd.DataFrame({"time": dates})
result_df = result_df.join(time_parameter_column)
result_df = result_df.set_index("time")
del result_df['test']

result_df.to_csv("sim_result.csv")

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,truncate_sheet=False, **to_excel_kwargs):
    from openpyxl import load_workbook
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
        
    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        writer.book = load_workbook(filename)

        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        if truncate_sheet and sheet_name in writer.book.sheetnames:
            idx = writer.book.sheetnames.index(sheet_name)
            writer.book.remove(writer.book.worksheets[idx])
            writer.book.create_sheet(sheet_name, idx)

        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        pass

    if startrow is None:
        startrow = 0

    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
    writer.save()
    
    
# Export to csv
def export_to_csv(path, team_name):
    
    df=pd.read_csv(path + f"/resultspace/{team_name}/sim_result.csv")
    names = df.columns[1:]
    wellnames = []
    for n in names:
        a=n.split(':')
        b=a[1]
        if b in wellnames:
            break
        else:
            wellnames.append(b)
    d = []
    for name in wellnames:
        d.append({'01_НГДУ':'RIENM Corp',    
                  '02_Месторождение': 'RIENM1',
                  '03_Скважина':name,         
                  '04_Тип_скважины':'верт',
                  '05_Пласт':"м1",         
                  '06_Dскв': 200,
                  '07_Dнкт':73,
                  '08_Н_вдп':2500,
                  '09_Удл':0,          
                  '10_Ндин':random.randint(200,500),
                  '11_СЭ':'ESP',
                  '12_Рзаб':df['WBHP:'+name].tail(20).mean(),        
                  '13_Qнефти':df['WOPR:'+name].tail(20).mean(),
                  '14_Qжидк':df['WLPR:'+name].tail(20).mean(),
                  '15_Обводненность':df['WWPR:'+name].tail(20).mean()/(1+df['WLPR:'+name].tail(20).mean()),
                  '16_Pзатр':15,
                  '17_ГФ':60,
                  '18_Тпл':104,
                  '19_Пл-ть_нефти':0.85,
                  '20_Пл-ть_воды':1 })
        df1=pd.DataFrame(d)
        append_df_to_excel(f"resultspace/{team_name}/201910_TR_1.xlsx", df1,sheet_name='TR', 
                            startrow=0, startcol=0, index=False, header=False)
        
